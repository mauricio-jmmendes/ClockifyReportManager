"""
Clockify Report Converter

This script converts a Clockify Detailed Time Report into a formatted Excel report
with both Summary and Detailed sheets. The Summary is automatically generated
by aggregating data from the Detailed report.

Usage:
    python clockify_report_converter.py [options]

Options:
    --rate <value>      Billable rate per hour in BRL (default: 50)
    --detailed <file>   Path to Detailed Excel file (auto-detected if not provided)
    --output <file>     Path for output Excel file (auto-generated if not provided)

Examples:
    1. Auto-detect Detailed file with default rate (50):
       python clockify_report_converter.py
    
    2. Auto-detect with custom rate:
       python clockify_report_converter.py --rate 250
    
    3. Manual file specification:
       python clockify_report_converter.py --rate 200 \\
           --detailed Clockify_Time_Report_Detailed_01_12_2025-26_12_2025.xlsx \\
           --output Time_Report_Output.xlsx
"""

import sys
import os
import glob
import argparse
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

# Default billable rate
DEFAULT_RATE = 50

# Get the directory where this script resides
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Color scheme matching the template
DARK_FILL = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
YELLOW_FONT = Font(color="FFFF00", bold=True)
YELLOW_FONT_NORMAL = Font(color="FFFF00", bold=False)
WHITE_FONT = Font(color="FFFFFF", bold=False, size=18)
WHITE_FONT_LARGE = Font(color="FFFFFF", bold=False, size=22)
HEADER_FONT = Font(color="FFFF00", bold=True, size=10)  # Yellow bold size 10 for headers
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
DATA_FONT = Font(size=10)  # Font size 10 for table data rows
TABLE_HEADER_FONT = Font(color="FFFF00", bold=True, size=10)  # Yellow bold for table headers
YELLOW_FONT_SIZE10 = Font(color="FFFF00", bold=True, size=10)  # Yellow bold size 10 for project rows


def parse_date_range_from_filename(filename: str) -> tuple[str, str]:
    """Extract date range from Clockify filename format."""
    import re
    # Pattern: DD_MM_YYYY-DD_MM_YYYY
    pattern = r'(\d{2}_\d{2}_\d{4})-(\d{2}_\d{2}_\d{4})'
    match = re.search(pattern, filename)
    if match:
        start = match.group(1).replace('_', '/')
        end = match.group(2).replace('_', '/')
        return start, end
    return None, None


def load_detailed_data(detailed_file: str) -> pd.DataFrame:
    """Load data from Clockify Detailed export file."""
    return pd.read_excel(detailed_file)


def decimal_to_time_str(decimal_hours: float) -> str:
    """Convert decimal hours to HH:MM:SS format."""
    total_seconds = int(round(decimal_hours * 3600))
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"


def time_str_to_decimal(time_str) -> float:
    """Convert HH:MM:SS string to decimal hours with full precision.
    
    This avoids rounding errors that occur when summing pre-rounded decimal values.
    """
    if pd.isna(time_str) or not time_str:
        return 0.0
    
    time_str = str(time_str)
    parts = time_str.split(':')
    
    if len(parts) == 3:
        try:
            hours = int(parts[0])
            minutes = int(parts[1])
            seconds = int(parts[2])
            return hours + minutes / 60 + seconds / 3600
        except ValueError:
            return 0.0
    elif len(parts) == 2:
        try:
            hours = int(parts[0])
            minutes = int(parts[1])
            return hours + minutes / 60
        except ValueError:
            return 0.0
    
    return 0.0


def build_summary_from_detailed(detailed_df: pd.DataFrame) -> list:
    """
    Build summary data by aggregating detailed report entries.
    
    Uses Duration (h) column (HH:MM:SS format) for precise calculations,
    avoiding rounding errors from pre-rounded decimal values.
    
    Returns a list of dictionaries with summary rows structured as:
    - Project header rows (with totals for each project)
    - Description rows (with totals for each description within a project)
    
    Note: Time (decimal) stores full precision; rounding happens at display time.
    """
    summary_rows = []
    
    # Group by Project (maintaining order from detailed report)
    for project in detailed_df['Project'].unique():
        project_data = detailed_df[detailed_df['Project'] == project]
        client = project_data['Client'].iloc[0]
        project_name = f"{project} - {client}" if pd.notna(client) else project
        
        # Calculate project totals using Duration (h) for full precision
        project_total_decimal = sum(
            time_str_to_decimal(t) for t in project_data['Duration (h)']
        )
        project_time_h = decimal_to_time_str(project_total_decimal)
        
        # Add project header row - store full precision, round at display
        summary_rows.append({
            'Project': project_name,
            'Description': None,
            'Time (h)': project_time_h,
            'Time (decimal)': project_total_decimal  # Full precision
        })
        
        # Group by Description within project (maintaining order)
        for description in project_data['Description'].unique():
            desc_data = project_data[project_data['Description'] == description]
            # Use Duration (h) for full precision
            desc_total_decimal = sum(
                time_str_to_decimal(t) for t in desc_data['Duration (h)']
            )
            desc_time_h = decimal_to_time_str(desc_total_decimal)
            
            summary_rows.append({
                'Project': None,
                'Description': description,
                'Time (h)': desc_time_h,
                'Time (decimal)': desc_total_decimal  # Full precision
            })
    
    return summary_rows


def create_summary_sheet(wb: Workbook, summary_data: list, date_range: tuple, rate: float = DEFAULT_RATE):
    """Create the Summary report sheet with proper formatting.
    
    Args:
        wb: Workbook to add the sheet to
        summary_data: List of dictionaries with summary rows (from build_summary_from_detailed)
        date_range: Tuple of (start_date, end_date) strings
        rate: Billable rate per hour
    """
    ws = wb.create_sheet("Summary report ")
    
    # Column widths
    col_widths = {'A': 39.14, 'B': 87.14, 'C': 16.29, 'D': 19.0, 'E': 21.86}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    # Row 1: Title with dark background - merge cells A1:D1 and center
    for col in range(1, 6):
        ws.cell(row=1, column=col).fill = DARK_FILL
    ws.merge_cells('A1:D1')
    ws['A1'] = "Summary report"
    ws['A1'].font = WHITE_FONT_LARGE
    ws['A1'].alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 30
    
    # Row 2: Empty with dark background
    for col in range(1, 6):
        ws.cell(row=2, column=col).fill = DARK_FILL
    
    # Row 3: Headers with dark background and yellow text
    headers = ['Project', 'Description', 'Time (h)', 'Time (decimal)', 'Amount (BRL)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = DARK_FILL
    
    # Row 4: Empty separator
    for col in range(1, 6):
        ws.cell(row=4, column=col).fill = DARK_FILL
    
    # Calculate total time decimal from project-level rows only
    total_time_decimal = sum(
        row['Time (decimal)'] for row in summary_data 
        if row['Project'] is not None
    )
    
    # Process summary data
    current_row = 5
    
    for row in summary_data:
        project = row['Project']
        description = row['Description']
        time_h = row['Time (h)']
        time_decimal = row['Time (decimal)']
        
        if project is not None:
            # This is a project header row - dark background, yellow text, size 10
            for col in range(1, 6):
                ws.cell(row=current_row, column=col).fill = DARK_FILL
            
            cell = ws.cell(row=current_row, column=1, value=project)
            cell.font = YELLOW_FONT_SIZE10
            
            cell = ws.cell(row=current_row, column=3, value=time_h)
            cell.font = YELLOW_FONT_SIZE10
            
            cell = ws.cell(row=current_row, column=4, value=time_decimal)
            cell.font = YELLOW_FONT_SIZE10
            cell.number_format = '#,##0.00'
            
            # Amount (BRL) = Time (decimal) * rate
            cell = ws.cell(row=current_row, column=5, value=f"=D{current_row}*{rate}")
            cell.font = YELLOW_FONT_SIZE10
            cell.number_format = '#,##0.00'
            
            current_row += 1
        elif description is not None:
            # This is a description row under a project - no fill, font size 10
            cell = ws.cell(row=current_row, column=2, value=description)
            cell.font = DATA_FONT
            cell = ws.cell(row=current_row, column=3, value=time_h)
            cell.font = DATA_FONT
            cell = ws.cell(row=current_row, column=4, value=time_decimal)
            cell.font = DATA_FONT
            cell.number_format = '#,##0.00'
            # Amount (BRL) = Time (decimal) * rate
            cell = ws.cell(row=current_row, column=5, value=f"=D{current_row}*{rate}")
            cell.font = DATA_FONT
            cell.number_format = '#,##0.00'
            current_row += 1
    
    # Total row - dark background, yellow text
    start_date, end_date = date_range
    total_label = f"Total ({start_date} - {end_date})" if start_date and end_date else "Total"
    
    # Apply dark background to entire total row
    for col in range(1, 6):
        ws.cell(row=current_row, column=col).fill = DARK_FILL
    
    cell = ws.cell(row=current_row, column=1, value=total_label)
    cell.font = YELLOW_FONT_SIZE10
    
    # Calculate total time in h:mm:ss format
    total_time_str = decimal_to_time_str(total_time_decimal)
    
    cell = ws.cell(row=current_row, column=3, value=total_time_str)
    cell.font = YELLOW_FONT_SIZE10
    
    cell = ws.cell(row=current_row, column=4, value=total_time_decimal)
    cell.font = YELLOW_FONT_SIZE10
    cell.number_format = '#,##0.00'
    
    # Formula for amount: time_decimal * rate
    cell = ws.cell(row=current_row, column=5, value=f"=D{current_row}*{rate}")
    cell.font = YELLOW_FONT_SIZE10
    cell.number_format = '#,##0.00'
    
    return ws


def create_detailed_sheet(wb: Workbook, detailed_df: pd.DataFrame, rate: float = DEFAULT_RATE):
    """Create the Detailed Report sheet with proper formatting and table."""
    ws = wb.create_sheet("Detailed Report")
    
    # Column widths
    col_widths = {
        'A': 17.29, 'B': 20.29, 'C': 87.14, 'D': 18.86, 'E': 11.29,
        'F': 12.57, 'G': 12.0, 'H': 13.14, 'I': 10.71, 'J': 12.43,
        'K': 9.43, 'L': 24.14, 'M': 30.57
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    # Row 1: Title and Total Amount with dark background
    for col in range(1, 14):
        ws.cell(row=1, column=col).fill = DARK_FILL
    
    # Merge and center "Detailed Report" title
    ws.merge_cells('C1:H1')
    ws['C1'] = "Detailed Report"
    ws['C1'].font = WHITE_FONT_LARGE
    ws['C1'].alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 30
    
    ws['L1'] = "Total Amount:"
    ws['L1'].font = YELLOW_FONT
    
    # Row 2: Empty with dark background
    for col in range(1, 14):
        ws.cell(row=2, column=col).fill = DARK_FILL
    
    # Row 3: Headers with dark background and yellow text
    headers = [
        'Project', 'Client', 'Description', 'User', 'Tags',
        'Start Date', 'Start Time', 'End Date', 'End Time',
        'Duration (h)', 'Duration (decimal)', 'Billable Rate (BRL)', 'Billable Amount (BRL)'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = TABLE_HEADER_FONT
        cell.fill = DARK_FILL
    
    # Process detailed data
    data_row = 4
    for idx, row in detailed_df.iterrows():
        # Column A: Project
        cell = ws.cell(row=data_row, column=1, value=row['Project'])
        cell.font = DATA_FONT
        
        # Column B: Client
        cell = ws.cell(row=data_row, column=2, value=row['Client'])
        cell.font = DATA_FONT
        
        # Column C: Description
        cell = ws.cell(row=data_row, column=3, value=row['Description'])
        cell.font = DATA_FONT
        
        # Column D: User
        cell = ws.cell(row=data_row, column=4, value=row['User'])
        cell.font = DATA_FONT
        
        # Column E: Tags (first tag if multiple)
        tags = row['Tags']
        cell = ws.cell(row=data_row, column=5)
        cell.font = DATA_FONT
        if pd.notna(tags):
            # Take first tag or clean up
            first_tag = str(tags).split(',')[0].strip()
            cell.value = first_tag
        
        # Column F: Start Date
        start_date = row['Start Date']
        cell = ws.cell(row=data_row, column=6)
        cell.font = DATA_FONT
        if pd.notna(start_date):
            if isinstance(start_date, str):
                try:
                    start_date = datetime.strptime(start_date, '%Y-%m-%d')
                except:
                    pass
            cell.value = start_date
            cell.number_format = 'dd/mm/yyyy'
        
        # Column G: Start Time
        start_time = row['Start Time']
        cell = ws.cell(row=data_row, column=7, value=start_time)
        cell.font = DATA_FONT
        cell.number_format = 'h:mm:ss'
        
        # Column H: End Date
        end_date = row['End Date']
        cell = ws.cell(row=data_row, column=8)
        cell.font = DATA_FONT
        if pd.notna(end_date):
            if isinstance(end_date, str):
                try:
                    end_date = datetime.strptime(end_date, '%Y-%m-%d')
                except:
                    pass
            cell.value = end_date
            cell.number_format = 'dd/mm/yyyy'
        
        # Column I: End Time
        end_time = row['End Time']
        cell = ws.cell(row=data_row, column=9, value=end_time)
        cell.font = DATA_FONT
        cell.number_format = 'h:mm:ss'
        
        # Column J: Duration (h)
        duration_h = row['Duration (h)']
        cell = ws.cell(row=data_row, column=10, value=duration_h)
        cell.font = DATA_FONT
        cell.number_format = 'h:mm:ss'
        
        # Column K: Duration (decimal) - use full precision from Duration (h)
        duration_decimal = time_str_to_decimal(duration_h)
        cell = ws.cell(row=data_row, column=11, value=duration_decimal)
        cell.font = DATA_FONT
        cell.number_format = '#,##0.00'
        
        # Column L: Billable Rate (BRL) - use the user-provided rate
        cell = ws.cell(row=data_row, column=12, value=rate)
        cell.font = DATA_FONT
        cell.number_format = '#,##0.00'
        
        # Column M: Billable Amount (BRL) - Formula
        cell = ws.cell(row=data_row, column=13, value=f"=L{data_row}*K{data_row}")
        cell.font = DATA_FONT
        cell.number_format = '#,##0.00'
        
        data_row += 1
    
    # Total Amount formula in M1 - Sum of all billable amounts
    last_data_row = data_row - 1
    ws['M1'] = f"=SUM(M4:M{last_data_row})"
    ws['M1'].font = YELLOW_FONT
    ws['M1'].number_format = '#,##0.00'
    
    # Create Table
    table_ref = f"A3:M{last_data_row}"
    table = Table(displayName="Table1", ref=table_ref)
    
    # Table Style
    style = TableStyleInfo(
        name="TableStyleMedium15",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)
    
    return ws


def convert_clockify_report(detailed_file: str, output_file: str, rate: float = DEFAULT_RATE):
    """Main function to convert Clockify Detailed report to the target format.
    
    The Summary sheet is automatically generated from the Detailed report data.
    """
    print(f"Loading data from:")
    print(f"  Detailed: {detailed_file}")
    print(f"  Billable Rate: {rate} BRL/hour")
    
    # Load detailed data
    detailed_df = load_detailed_data(detailed_file)
    
    print(f"\nDetailed rows: {len(detailed_df)}")
    
    # Build summary from detailed data
    summary_data = build_summary_from_detailed(detailed_df)
    print(f"Summary rows generated: {len(summary_data)}")
    
    # Extract date range from filename
    date_range = parse_date_range_from_filename(detailed_file)
    
    print(f"Date range: {date_range[0]} - {date_range[1]}")
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    
    # Create sheets
    create_summary_sheet(wb, summary_data, date_range, rate)
    create_detailed_sheet(wb, detailed_df, rate)
    
    # Remove default sheet
    wb.remove(default_sheet)
    
    # Save workbook
    wb.save(output_file)
    print(f"\nOutput saved to: {output_file}")


def find_detailed_file():
    """
    Search for Clockify Detailed export file in the script's directory.
    Returns the path to the most recent file, or None if not found.
    """
    # Search for Detailed file
    detailed_pattern = os.path.join(SCRIPT_DIR, "Clockify_Time_Report_Detailed_*.xlsx")
    detailed_files = glob.glob(detailed_pattern)
    
    if not detailed_files:
        print(f"Error: No Detailed file found matching pattern: Clockify_Time_Report_Detailed_*.xlsx")
        print(f"Searched in: {SCRIPT_DIR}")
        return None
    
    # Use the most recent file if multiple matches (sorted by modification time)
    detailed_file = max(detailed_files, key=os.path.getmtime)
    
    return detailed_file


def main():
    parser = argparse.ArgumentParser(
        description='Convert Clockify Detailed Report to formatted Excel with Summary and Detailed sheets.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__
    )
    parser.add_argument('--rate', type=float, default=DEFAULT_RATE,
                        help=f'Billable rate per hour in BRL (default: {DEFAULT_RATE})')
    parser.add_argument('--detailed', type=str, default=None,
                        help='Path to Detailed Excel file (auto-detected if not provided)')
    parser.add_argument('--output', type=str, default=None,
                        help='Path for output Excel file (auto-generated if not provided)')
    
    args = parser.parse_args()
    
    # Determine input file
    if args.detailed:
        detailed_file = args.detailed
    else:
        # Auto-detect file in script directory
        detailed_file = find_detailed_file()
        
        if detailed_file is None:
            sys.exit(1)
    
    # Determine output file
    if args.output:
        output_file = args.output
    else:
        # Generate output filename based on input file date range
        date_range = parse_date_range_from_filename(detailed_file)
        if date_range[0] and date_range[1]:
            output_name = f"Time_Report_Generated_{date_range[0].replace('/', '_')}-{date_range[1].replace('/', '_')}.xlsx"
        else:
            output_name = "Time_Report_Generated.xlsx"
        output_file = os.path.join(SCRIPT_DIR, output_name)
    
    convert_clockify_report(detailed_file, output_file, args.rate)


if __name__ == "__main__":
    main()

