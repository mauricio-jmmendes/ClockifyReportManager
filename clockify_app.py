"""
Clockify Report Converter - Desktop App

A modern desktop application to convert Clockify Detailed Time Reports
into formatted Excel reports. The Summary sheet is automatically generated
by aggregating data from the Detailed report.
"""

import os
import sys
import glob
import re
import threading
from datetime import datetime
from pathlib import Path

import customtkinter as ctk
from tkinter import filedialog, messagebox

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

# ============================================================================
# CONFIGURATION
# ============================================================================

DEFAULT_RATE = 50
APP_NAME = "Clockify Report Converter"
VERSION = "1.0.0"

# Get the directory where this app/script resides
if getattr(sys, 'frozen', False):
    # Running as compiled executable
    APP_DIR = os.path.dirname(sys.executable)
else:
    # Running as script
    APP_DIR = os.path.dirname(os.path.abspath(__file__))

# Color scheme
ACCENT_COLOR = "#00D4AA"  # Teal/mint accent
ACCENT_HOVER = "#00B894"
DARK_BG = "#1a1a2e"
CARD_BG = "#16213e"
INPUT_BG = "#0f3460"
TEXT_PRIMARY = "#ffffff"
TEXT_SECONDARY = "#a0a0a0"
SUCCESS_COLOR = "#00D4AA"
ERROR_COLOR = "#ff6b6b"

# Excel styling
DARK_FILL = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
YELLOW_FONT = Font(color="FFFF00", bold=True)
WHITE_FONT = Font(color="FFFFFF", bold=False, size=18)
WHITE_FONT_LARGE = Font(color="FFFFFF", bold=False, size=22)
HEADER_FONT = Font(color="FFFF00", bold=True, size=10)
DATA_FONT = Font(size=10)
TABLE_HEADER_FONT = Font(color="FFFF00", bold=True, size=10)
YELLOW_FONT_SIZE10 = Font(color="FFFF00", bold=True, size=10)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')

# ============================================================================
# CORE CONVERSION LOGIC
# ============================================================================

def parse_date_range_from_filename(filename: str) -> tuple[str, str]:
    """Extract date range from Clockify filename format."""
    pattern = r'(\d{2}_\d{2}_\d{4})-(\d{2}_\d{2}_\d{4})'
    match = re.search(pattern, filename)
    if match:
        start = match.group(1).replace('_', '/')
        end = match.group(2).replace('_', '/')
        return start, end
    return None, None


def load_detailed_data(detailed_file: str) -> pd.DataFrame:
    """Load data from Clockify Detailed export file."""
    return pd.read_excel(detailed_file)


def decimal_to_time_str(decimal_hours: float) -> str:
    """Convert decimal hours to HH:MM:SS format."""
    total_seconds = int(round(decimal_hours * 3600))
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"


def time_str_to_decimal(time_str) -> float:
    """Convert HH:MM:SS string to decimal hours with full precision.
    
    This avoids rounding errors that occur when summing pre-rounded decimal values.
    """
    if pd.isna(time_str) or not time_str:
        return 0.0
    
    time_str = str(time_str)
    parts = time_str.split(':')
    
    if len(parts) == 3:
        try:
            hours = int(parts[0])
            minutes = int(parts[1])
            seconds = int(parts[2])
            return hours + minutes / 60 + seconds / 3600
        except ValueError:
            return 0.0
    elif len(parts) == 2:
        try:
            hours = int(parts[0])
            minutes = int(parts[1])
            return hours + minutes / 60
        except ValueError:
            return 0.0
    
    return 0.0


def build_summary_from_detailed(detailed_df: pd.DataFrame) -> list:
    """
    Build summary data by aggregating detailed report entries.
    
    Uses Duration (h) column (HH:MM:SS format) for precise calculations,
    avoiding rounding errors from pre-rounded decimal values.
    
    Returns a list of dictionaries with summary rows structured as:
    - Project header rows (with totals for each project)
    - Description rows (with totals for each description within a project)
    
    Note: Time (decimal) stores full precision; rounding happens at display time.
    """
    summary_rows = []
    
    # Group by Project (maintaining order from detailed report)
    for project in detailed_df['Project'].unique():
        project_data = detailed_df[detailed_df['Project'] == project]
        client = project_data['Client'].iloc[0]
        project_name = f"{project} - {client}" if pd.notna(client) else project
        
        # Calculate project totals using Duration (h) for full precision
        project_total_decimal = sum(
            time_str_to_decimal(t) for t in project_data['Duration (h)']
        )
        project_time_h = decimal_to_time_str(project_total_decimal)
        
        # Add project header row - store full precision, round at display
        summary_rows.append({
            'Project': project_name,
            'Description': None,
            'Time (h)': project_time_h,
            'Time (decimal)': project_total_decimal  # Full precision
        })
        
        # Group by Description within project (maintaining order)
        for description in project_data['Description'].unique():
            desc_data = project_data[project_data['Description'] == description]
            # Use Duration (h) for full precision
            desc_total_decimal = sum(
                time_str_to_decimal(t) for t in desc_data['Duration (h)']
            )
            desc_time_h = decimal_to_time_str(desc_total_decimal)
            
            summary_rows.append({
                'Project': None,
                'Description': description,
                'Time (h)': desc_time_h,
                'Time (decimal)': desc_total_decimal  # Full precision
            })
    
    return summary_rows


def create_summary_sheet(wb: Workbook, summary_data: list, date_range: tuple, rate: float = DEFAULT_RATE):
    """Create the Summary report sheet with proper formatting.
    
    Args:
        wb: Workbook to add the sheet to
        summary_data: List of dictionaries with summary rows (from build_summary_from_detailed)
        date_range: Tuple of (start_date, end_date) strings
        rate: Billable rate per hour
    """
    ws = wb.create_sheet("Summary report ")
    
    col_widths = {'A': 39.14, 'B': 87.14, 'C': 16.29, 'D': 19.0, 'E': 21.86}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    for col in range(1, 6):
        ws.cell(row=1, column=col).fill = DARK_FILL
    ws.merge_cells('A1:D1')
    ws['A1'] = "Summary report"
    ws['A1'].font = WHITE_FONT_LARGE
    ws['A1'].alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 30
    
    for col in range(1, 6):
        ws.cell(row=2, column=col).fill = DARK_FILL
    
    headers = ['Project', 'Description', 'Time (h)', 'Time (decimal)', 'Amount (BRL)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = DARK_FILL
    
    for col in range(1, 6):
        ws.cell(row=4, column=col).fill = DARK_FILL
    
    # Calculate total time decimal from project-level rows only
    total_time_decimal = sum(
        row['Time (decimal)'] for row in summary_data 
        if row['Project'] is not None
    )
    
    current_row = 5
    
    for row in summary_data:
        project = row['Project']
        description = row['Description']
        time_h = row['Time (h)']
        time_decimal = row['Time (decimal)']
        
        if project is not None:
            for col in range(1, 6):
                ws.cell(row=current_row, column=col).fill = DARK_FILL
            
            cell = ws.cell(row=current_row, column=1, value=project)
            cell.font = YELLOW_FONT_SIZE10
            
            cell = ws.cell(row=current_row, column=3, value=time_h)
            cell.font = YELLOW_FONT_SIZE10
            
            cell = ws.cell(row=current_row, column=4, value=time_decimal)
            cell.font = YELLOW_FONT_SIZE10
            cell.number_format = '#,##0.00'
            
            cell = ws.cell(row=current_row, column=5, value=f"=D{current_row}*{rate}")
            cell.font = YELLOW_FONT_SIZE10
            cell.number_format = '#,##0.00'
            
            current_row += 1
        elif description is not None:
            cell = ws.cell(row=current_row, column=2, value=description)
            cell.font = DATA_FONT
            cell = ws.cell(row=current_row, column=3, value=time_h)
            cell.font = DATA_FONT
            cell = ws.cell(row=current_row, column=4, value=time_decimal)
            cell.font = DATA_FONT
            cell.number_format = '#,##0.00'
            cell = ws.cell(row=current_row, column=5, value=f"=D{current_row}*{rate}")
            cell.font = DATA_FONT
            cell.number_format = '#,##0.00'
            current_row += 1
    
    start_date, end_date = date_range
    total_label = f"Total ({start_date} - {end_date})" if start_date and end_date else "Total"
    
    for col in range(1, 6):
        ws.cell(row=current_row, column=col).fill = DARK_FILL
    
    cell = ws.cell(row=current_row, column=1, value=total_label)
    cell.font = YELLOW_FONT_SIZE10
    
    total_time_str = decimal_to_time_str(total_time_decimal)
    
    cell = ws.cell(row=current_row, column=3, value=total_time_str)
    cell.font = YELLOW_FONT_SIZE10
    
    cell = ws.cell(row=current_row, column=4, value=total_time_decimal)
    cell.font = YELLOW_FONT_SIZE10
    cell.number_format = '#,##0.00'
    
    cell = ws.cell(row=current_row, column=5, value=f"=D{current_row}*{rate}")
    cell.font = YELLOW_FONT_SIZE10
    cell.number_format = '#,##0.00'
    
    return ws


def create_detailed_sheet(wb: Workbook, detailed_df: pd.DataFrame, rate: float = DEFAULT_RATE):
    """Create the Detailed Report sheet with proper formatting and table."""
    ws = wb.create_sheet("Detailed Report")
    
    col_widths = {
        'A': 17.29, 'B': 20.29, 'C': 87.14, 'D': 18.86, 'E': 11.29,
        'F': 12.57, 'G': 12.0, 'H': 13.14, 'I': 10.71, 'J': 12.43,
        'K': 9.43, 'L': 24.14, 'M': 30.57
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    for col in range(1, 14):
        ws.cell(row=1, column=col).fill = DARK_FILL
    
    ws.merge_cells('C1:H1')
    ws['C1'] = "Detailed Report"
    ws['C1'].font = WHITE_FONT_LARGE
    ws['C1'].alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 30
    
    ws['L1'] = "Total Amount:"
    ws['L1'].font = YELLOW_FONT
    
    for col in range(1, 14):
        ws.cell(row=2, column=col).fill = DARK_FILL
    
    headers = [
        'Project', 'Client', 'Description', 'User', 'Tags',
        'Start Date', 'Start Time', 'End Date', 'End Time',
        'Duration (h)', 'Duration (decimal)', 'Billable Rate (BRL)', 'Billable Amount (BRL)'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = TABLE_HEADER_FONT
        cell.fill = DARK_FILL
    
    data_row = 4
    for idx, row in detailed_df.iterrows():
        cell = ws.cell(row=data_row, column=1, value=row['Project'])
        cell.font = DATA_FONT
        
        cell = ws.cell(row=data_row, column=2, value=row['Client'])
        cell.font = DATA_FONT
        
        cell = ws.cell(row=data_row, column=3, value=row['Description'])
        cell.font = DATA_FONT
        
        cell = ws.cell(row=data_row, column=4, value=row['User'])
        cell.font = DATA_FONT
        
        tags = row['Tags']
        cell = ws.cell(row=data_row, column=5)
        cell.font = DATA_FONT
        if pd.notna(tags):
            first_tag = str(tags).split(',')[0].strip()
            cell.value = first_tag
        
        start_date = row['Start Date']
        cell = ws.cell(row=data_row, column=6)
        cell.font = DATA_FONT
        if pd.notna(start_date):
            if isinstance(start_date, str):
                try:
                    start_date = datetime.strptime(start_date, '%Y-%m-%d')
                except:
                    pass
            cell.value = start_date
            cell.number_format = 'dd/mm/yyyy'
        
        start_time = row['Start Time']
        cell = ws.cell(row=data_row, column=7, value=start_time)
        cell.font = DATA_FONT
        cell.number_format = 'h:mm:ss'
        
        end_date = row['End Date']
        cell = ws.cell(row=data_row, column=8)
        cell.font = DATA_FONT
        if pd.notna(end_date):
            if isinstance(end_date, str):
                try:
                    end_date = datetime.strptime(end_date, '%Y-%m-%d')
                except:
                    pass
            cell.value = end_date
            cell.number_format = 'dd/mm/yyyy'
        
        end_time = row['End Time']
        cell = ws.cell(row=data_row, column=9, value=end_time)
        cell.font = DATA_FONT
        cell.number_format = 'h:mm:ss'
        
        duration_h = row['Duration (h)']
        cell = ws.cell(row=data_row, column=10, value=duration_h)
        cell.font = DATA_FONT
        cell.number_format = 'h:mm:ss'
        
        # Use full precision from Duration (h) instead of pre-rounded decimal
        duration_decimal = time_str_to_decimal(duration_h)
        cell = ws.cell(row=data_row, column=11, value=duration_decimal)
        cell.font = DATA_FONT
        cell.number_format = '#,##0.00'
        
        cell = ws.cell(row=data_row, column=12, value=rate)
        cell.font = DATA_FONT
        cell.number_format = '#,##0.00'
        
        cell = ws.cell(row=data_row, column=13, value=f"=L{data_row}*K{data_row}")
        cell.font = DATA_FONT
        cell.number_format = '#,##0.00'
        
        data_row += 1
    
    last_data_row = data_row - 1
    ws['M1'] = f"=SUM(M4:M{last_data_row})"
    ws['M1'].font = YELLOW_FONT
    ws['M1'].number_format = '#,##0.00'
    
    table_ref = f"A3:M{last_data_row}"
    table = Table(displayName="Table1", ref=table_ref)
    
    style = TableStyleInfo(
        name="TableStyleMedium15",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)
    
    return ws


def convert_clockify_report(detailed_file: str, output_file: str, rate: float = DEFAULT_RATE):
    """Main function to convert Clockify Detailed report to the target format.
    
    The Summary sheet is automatically generated from the Detailed report data.
    
    Returns:
        Tuple of (summary_rows_count, detailed_rows_count)
    """
    # Load detailed data
    detailed_df = load_detailed_data(detailed_file)
    
    # Build summary from detailed data
    summary_data = build_summary_from_detailed(detailed_df)
    
    # Extract date range from filename
    date_range = parse_date_range_from_filename(detailed_file)
    
    wb = Workbook()
    default_sheet = wb.active
    
    create_summary_sheet(wb, summary_data, date_range, rate)
    create_detailed_sheet(wb, detailed_df, rate)
    
    wb.remove(default_sheet)
    wb.save(output_file)
    
    return len(summary_data), len(detailed_df)


def find_detailed_file():
    """
    Search for Clockify Detailed export file in the app's directory.
    Returns the path to the most recent file, or None if not found.
    """
    # Search for Detailed file
    detailed_pattern = os.path.join(APP_DIR, "Clockify_Time_Report_Detailed_*.xlsx")
    detailed_files = glob.glob(detailed_pattern)
    
    if not detailed_files:
        return None
    
    # Use the most recent file if multiple matches
    return max(detailed_files, key=os.path.getmtime)


# ============================================================================
# GUI APPLICATION
# ============================================================================

class FileDropFrame(ctk.CTkFrame):
    """A styled frame for file selection with visual feedback."""
    
    def __init__(self, master, label_text, file_type, **kwargs):
        super().__init__(master, **kwargs)
        
        self.file_path = None
        self.file_type = file_type
        
        self.configure(
            fg_color=CARD_BG,
            corner_radius=12,
            border_width=2,
            border_color=INPUT_BG
        )
        
        # Icon/emoji
        icon = "📋"
        
        self.icon_label = ctk.CTkLabel(
            self,
            text=icon,
            font=ctk.CTkFont(size=32)
        )
        self.icon_label.pack(pady=(20, 5))
        
        self.title_label = ctk.CTkLabel(
            self,
            text=label_text,
            font=ctk.CTkFont(family="Segoe UI", size=14, weight="bold"),
            text_color=TEXT_PRIMARY
        )
        self.title_label.pack(pady=(0, 5))
        
        self.subtitle_label = ctk.CTkLabel(
            self,
            text="Summary will be auto-generated",
            font=ctk.CTkFont(family="Segoe UI", size=10),
            text_color=ACCENT_COLOR
        )
        self.subtitle_label.pack(pady=(0, 5))
        
        self.file_label = ctk.CTkLabel(
            self,
            text="No file selected",
            font=ctk.CTkFont(family="Segoe UI", size=11),
            text_color=TEXT_SECONDARY,
            wraplength=400
        )
        self.file_label.pack(pady=(0, 10))
        
        self.browse_btn = ctk.CTkButton(
            self,
            text="Browse",
            font=ctk.CTkFont(family="Segoe UI", size=12),
            fg_color=INPUT_BG,
            hover_color=ACCENT_HOVER,
            corner_radius=8,
            height=32,
            width=120,
            command=self.browse_file
        )
        self.browse_btn.pack(pady=(0, 20))
    
    def browse_file(self):
        file_path = filedialog.askopenfilename(
            title=f"Select {self.file_type} File",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if file_path:
            self.set_file(file_path)
    
    def set_file(self, file_path):
        self.file_path = file_path
        filename = os.path.basename(file_path)
        self.file_label.configure(text=filename, text_color=SUCCESS_COLOR)
        self.configure(border_color=SUCCESS_COLOR)
    
    def get_file(self):
        return self.file_path
    
    def reset(self):
        self.file_path = None
        self.file_label.configure(text="No file selected", text_color=TEXT_SECONDARY)
        self.configure(border_color=INPUT_BG)


class ClockifyApp(ctk.CTk):
    """Main application window."""
    
    def __init__(self):
        super().__init__()
        
        # Window setup
        self.title(APP_NAME)
        self.geometry("700x720")
        self.minsize(600, 680)
        self.configure(fg_color=DARK_BG)
        
        # Center window on screen
        self.update_idletasks()
        x = (self.winfo_screenwidth() - 700) // 2
        y = (self.winfo_screenheight() - 720) // 2
        self.geometry(f"700x720+{x}+{y}")
        
        self._create_widgets()
    
    def _create_widgets(self):
        # Main container with padding
        main_frame = ctk.CTkFrame(self, fg_color="transparent")
        main_frame.pack(fill="both", expand=True, padx=30, pady=20)
        
        # Header
        header_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        header_frame.pack(fill="x", pady=(0, 20))
        
        title_label = ctk.CTkLabel(
            header_frame,
            text="⏱️ Clockify Report Converter",
            font=ctk.CTkFont(family="Segoe UI", size=28, weight="bold"),
            text_color=TEXT_PRIMARY
        )
        title_label.pack(anchor="w")
        
        subtitle_label = ctk.CTkLabel(
            header_frame,
            text="Transform your Clockify exports into beautiful formatted reports",
            font=ctk.CTkFont(family="Segoe UI", size=13),
            text_color=TEXT_SECONDARY
        )
        subtitle_label.pack(anchor="w", pady=(5, 0))
        
        # File selection section
        files_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        files_frame.pack(fill="x", pady=(10, 20))
        
        files_label = ctk.CTkLabel(
            files_frame,
            text="INPUT FILE",
            font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"),
            text_color=ACCENT_COLOR
        )
        files_label.pack(anchor="w", pady=(0, 10))
        
        # Single file drop zone for Detailed report
        self.detailed_frame = FileDropFrame(
            files_frame,
            label_text="Clockify Detailed Report",
            file_type="Detailed",
            height=160
        )
        self.detailed_frame.pack(fill="x")
        
        # Settings section
        settings_frame = ctk.CTkFrame(main_frame, fg_color=CARD_BG, corner_radius=12)
        settings_frame.pack(fill="x", pady=(0, 20))
        
        settings_inner = ctk.CTkFrame(settings_frame, fg_color="transparent")
        settings_inner.pack(fill="x", padx=20, pady=20)
        
        settings_label = ctk.CTkLabel(
            settings_inner,
            text="SETTINGS",
            font=ctk.CTkFont(family="Segoe UI", size=11, weight="bold"),
            text_color=ACCENT_COLOR
        )
        settings_label.pack(anchor="w", pady=(0, 15))
        
        # User name row
        name_row = ctk.CTkFrame(settings_inner, fg_color="transparent")
        name_row.pack(fill="x", pady=(0, 10))
        
        name_label = ctk.CTkLabel(
            name_row,
            text="User Name",
            font=ctk.CTkFont(family="Segoe UI", size=13),
            text_color=TEXT_PRIMARY
        )
        name_label.pack(side="left")
        
        self.name_entry = ctk.CTkEntry(
            name_row,
            width=200,
            height=38,
            font=ctk.CTkFont(family="Segoe UI", size=14),
            fg_color=INPUT_BG,
            border_color=INPUT_BG,
            corner_radius=8,
            placeholder_text="e.g. Mauricio Mendes"
        )
        self.name_entry.pack(side="right")
        
        # Rate row
        rate_row = ctk.CTkFrame(settings_inner, fg_color="transparent")
        rate_row.pack(fill="x", pady=(0, 10))
        
        rate_label = ctk.CTkLabel(
            rate_row,
            text="Billable Rate (BRL/hour)",
            font=ctk.CTkFont(family="Segoe UI", size=13),
            text_color=TEXT_PRIMARY
        )
        rate_label.pack(side="left")
        
        self.rate_entry = ctk.CTkEntry(
            rate_row,
            width=120,
            height=38,
            font=ctk.CTkFont(family="Segoe UI", size=14),
            fg_color=INPUT_BG,
            border_color=INPUT_BG,
            corner_radius=8,
            justify="center"
        )
        self.rate_entry.pack(side="right")
        self.rate_entry.insert(0, str(DEFAULT_RATE))
        
        # Output folder row
        output_row = ctk.CTkFrame(settings_inner, fg_color="transparent")
        output_row.pack(fill="x")
        
        output_label = ctk.CTkLabel(
            output_row,
            text="Output Folder",
            font=ctk.CTkFont(family="Segoe UI", size=13),
            text_color=TEXT_PRIMARY
        )
        output_label.pack(side="left")
        
        output_right = ctk.CTkFrame(output_row, fg_color="transparent")
        output_right.pack(side="right")
        
        self.output_entry = ctk.CTkEntry(
            output_right,
            width=280,
            height=38,
            font=ctk.CTkFont(family="Segoe UI", size=12),
            fg_color=INPUT_BG,
            border_color=INPUT_BG,
            corner_radius=8
        )
        self.output_entry.pack(side="left", padx=(0, 8))
        self.output_entry.insert(0, APP_DIR)
        
        self.output_browse_btn = ctk.CTkButton(
            output_right,
            text="📁",
            font=ctk.CTkFont(size=16),
            fg_color=INPUT_BG,
            hover_color=ACCENT_HOVER,
            corner_radius=8,
            width=38,
            height=38,
            command=self.browse_output_folder
        )
        self.output_browse_btn.pack(side="left")
        
        # Convert button
        self.convert_btn = ctk.CTkButton(
            main_frame,
            text="Convert Reports",
            font=ctk.CTkFont(family="Segoe UI", size=16, weight="bold"),
            fg_color=ACCENT_COLOR,
            hover_color=ACCENT_HOVER,
            text_color="#000000",
            corner_radius=10,
            height=50,
            command=self.start_conversion
        )
        self.convert_btn.pack(fill="x", pady=(0, 15))
        
        # Progress bar (hidden by default)
        self.progress_bar = ctk.CTkProgressBar(
            main_frame,
            fg_color=INPUT_BG,
            progress_color=ACCENT_COLOR,
            corner_radius=5,
            height=6
        )
        self.progress_bar.set(0)
        
        # Status section
        status_frame = ctk.CTkFrame(main_frame, fg_color=CARD_BG, corner_radius=12)
        status_frame.pack(fill="x", pady=(0, 10))
        
        self.status_label = ctk.CTkLabel(
            status_frame,
            text="Ready to convert",
            font=ctk.CTkFont(family="Segoe UI", size=12),
            text_color=TEXT_SECONDARY,
            anchor="w"
        )
        self.status_label.pack(fill="x", padx=20, pady=15)
        
        # Footer
        footer_label = ctk.CTkLabel(
            main_frame,
            text=f"v{VERSION}",
            font=ctk.CTkFont(family="Segoe UI", size=10),
            text_color=TEXT_SECONDARY
        )
        footer_label.pack(side="bottom", pady=(10, 0))
        
        # Auto-detect input files
        self._auto_detect_files()
    
    def _auto_detect_files(self):
        """Try to find Clockify Detailed file in the app directory and pre-populate."""
        detailed_file = find_detailed_file()
        
        if detailed_file:
            self.detailed_frame.set_file(detailed_file)
    
    def browse_output_folder(self):
        """Open folder browser for output directory."""
        folder = filedialog.askdirectory(
            title="Select Output Folder",
            initialdir=self.output_entry.get() or APP_DIR
        )
        if folder:
            self.output_entry.delete(0, "end")
            self.output_entry.insert(0, folder)
    
    def set_status(self, message, is_error=False, is_success=False):
        """Update status message with appropriate styling."""
        color = ERROR_COLOR if is_error else (SUCCESS_COLOR if is_success else TEXT_SECONDARY)
        self.status_label.configure(text=message, text_color=color)
    
    def start_conversion(self):
        """Validate inputs and start conversion in background thread."""
        # Validate file
        detailed_file = self.detailed_frame.get_file()
        
        if not detailed_file:
            self.set_status("Please select a Detailed report file", is_error=True)
            return
        
        # Validate rate
        try:
            rate = float(self.rate_entry.get())
            if rate <= 0:
                raise ValueError("Rate must be positive")
        except ValueError:
            self.set_status("Please enter a valid billable rate", is_error=True)
            return
        
        # Get user name
        user_name = self.name_entry.get().strip()
        if not user_name:
            self.set_status("Please enter a user name", is_error=True)
            return
        
        # Get output folder
        output_folder = self.output_entry.get().strip()
        if not output_folder:
            output_folder = APP_DIR
        if not os.path.isdir(output_folder):
            self.set_status("Output folder does not exist", is_error=True)
            return
        
        # Disable UI during conversion
        self.convert_btn.configure(state="disabled", text="Converting...")
        self.progress_bar.pack(fill="x", pady=(0, 15), before=self.status_label.master)
        self.progress_bar.start()
        self.set_status("Converting reports...")
        
        # Run conversion in background thread
        thread = threading.Thread(
            target=self._run_conversion,
            args=(detailed_file, rate, user_name, output_folder),
            daemon=True
        )
        thread.start()
    
    def _generate_output_path(self, detailed_file, user_name, output_folder):
        """Generate the output file path based on user name and date range."""
        user_name_formatted = user_name.replace(' ', '_')
        
        date_range = parse_date_range_from_filename(detailed_file)
        if date_range[0] and date_range[1]:
            base_name = f"{user_name_formatted}_Time_Report_{date_range[0].replace('/', '_')}-{date_range[1].replace('/', '_')}"
        else:
            base_name = f"{user_name_formatted}_Time_Report"
        
        return output_folder, base_name
    
    def _get_unique_filename(self, output_folder, base_name):
        """Get a unique filename by appending (1), (2), etc. if file exists."""
        output_file = os.path.join(output_folder, f"{base_name}.xlsx")
        
        if not os.path.exists(output_file):
            return output_file
        
        # Find next available number
        counter = 1
        while True:
            output_file = os.path.join(output_folder, f"{base_name}({counter}).xlsx")
            if not os.path.exists(output_file):
                return output_file
            counter += 1
    
    def _run_conversion(self, detailed_file, rate, user_name, output_folder):
        """Run the actual conversion (called from background thread)."""
        try:
            # Generate base output path
            folder, base_name = self._generate_output_path(detailed_file, user_name, output_folder)
            output_file = os.path.join(folder, f"{base_name}.xlsx")
            
            # Check if file exists - need to handle on main thread for dialog
            if os.path.exists(output_file):
                # Schedule dialog on main thread and wait for response
                self.after(0, lambda: self._handle_file_exists(
                    detailed_file, rate, output_file, folder, base_name
                ))
                return
            
            # Perform conversion
            summary_rows, detailed_rows = convert_clockify_report(
                detailed_file, output_file, rate
            )
            
            # Update UI on main thread
            self.after(0, lambda: self._conversion_complete(output_file, summary_rows, detailed_rows))
            
        except Exception as e:
            error_msg = str(e)
            self.after(0, lambda msg=error_msg: self._conversion_error(msg))
    
    def _handle_file_exists(self, detailed_file, rate, output_file, folder, base_name):
        """Handle case when output file already exists."""
        # Ask user what to do
        response = messagebox.askyesnocancel(
            "File Already Exists",
            f"The file already exists:\n\n{os.path.basename(output_file)}\n\n"
            f"Do you want to overwrite it?\n\n"
            f"• Yes - Overwrite the existing file\n"
            f"• No - Create a new file with a number suffix\n"
            f"• Cancel - Abort the conversion"
        )
        
        if response is None:
            # User clicked Cancel
            self.progress_bar.stop()
            self.progress_bar.pack_forget()
            self.convert_btn.configure(state="normal", text="Convert Reports")
            self.set_status("Conversion cancelled", is_error=False)
            return
        
        if response:
            # User clicked Yes - overwrite
            final_output = output_file
        else:
            # User clicked No - create numbered file
            final_output = self._get_unique_filename(folder, base_name)
        
        # Run conversion in background thread
        thread = threading.Thread(
            target=self._do_conversion,
            args=(detailed_file, rate, final_output),
            daemon=True
        )
        thread.start()
    
    def _do_conversion(self, detailed_file, rate, output_file):
        """Perform the actual file conversion."""
        try:
            summary_rows, detailed_rows = convert_clockify_report(
                detailed_file, output_file, rate
            )
            self.after(0, lambda: self._conversion_complete(output_file, summary_rows, detailed_rows))
        except Exception as e:
            error_msg = str(e)
            self.after(0, lambda msg=error_msg: self._conversion_error(msg))
    
    def _conversion_complete(self, output_file, summary_rows, detailed_rows):
        """Handle successful conversion (called on main thread)."""
        self.progress_bar.stop()
        self.progress_bar.pack_forget()
        self.convert_btn.configure(state="normal", text="Convert Reports")
        
        self.set_status(f"✓ Report saved: {os.path.basename(output_file)}", is_success=True)
        
        # Ask if user wants to open the file
        if messagebox.askyesno(
            "Conversion Complete",
            f"Report generated successfully!\n\n"
            f"Summary: {summary_rows} rows\n"
            f"Detailed: {detailed_rows} rows\n\n"
            f"Saved to:\n{output_file}\n\n"
            f"Would you like to open the file?"
        ):
            os.startfile(output_file)
    
    def _conversion_error(self, error_message):
        """Handle conversion error (called on main thread)."""
        self.progress_bar.stop()
        self.progress_bar.pack_forget()
        self.convert_btn.configure(state="normal", text="Convert Reports")
        
        self.set_status(f"Error: {error_message}", is_error=True)
        messagebox.showerror("Conversion Error", f"An error occurred:\n\n{error_message}")


# ============================================================================
# ENTRY POINT
# ============================================================================

def main():
    # Set appearance mode and theme
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("dark-blue")
    
    app = ClockifyApp()
    app.mainloop()


if __name__ == "__main__":
    main()

